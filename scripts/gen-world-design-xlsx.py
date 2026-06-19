#!/usr/bin/env python3
"""Generate docs/td-world-design.xlsx for game design fill-in."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "td-world-design.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F5A48")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FONT = Font(italic=True, color="666666")


def set_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def auto_width(ws, max_col: int, min_w=10, max_w=36) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_w
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[letter].width = best


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # --- 说明 ---
    ws0 = wb.create_sheet("填写说明", 0)
    notes = [
        "小人物 · 怪物区域 & 装备养成设计表",
        "",
        "1. 在下方各工作表填写中文名称和数值，留空的我可以按公式补全。",
        "2. 稀有度：普通 / 高级 / 稀有 / 传说 / 特制（特制打怪极低掉落，偶尔商城）。",
        "3. 填好后保存，发给我或告诉我「表填好了」。",
        "4. 「怪物区域」= 原爬塔，将改名为一级、二级…区域。",
        "",
        "优先填写：怪物区域、装备实例 两张表。",
    ]
    for i, line in enumerate(notes, 1):
        ws0.cell(row=i, column=1, value=line)
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 80

    # --- 主角等级 ---
    ws1 = wb.create_sheet("主角等级")
    h1 = ["等级", "所需累计经验", "解锁怪物区域最高级", "可穿戴装备等级上限", "备注"]
    set_header(ws1, 1, h1)
    samples1 = [
        (1, 0, 1, 1, "初始"),
        (2, 100, 1, 2, ""),
        (3, 250, 2, 3, ""),
        (5, 800, 3, 5, ""),
        (10, 3000, 5, 10, ""),
        (20, 15000, 10, 15, ""),
        (30, 40000, 15, 20, ""),
        (50, 120000, 20, 30, "满级"),
    ]
    for r, row in enumerate(samples1, 2):
        for c, v in enumerate(row, 1):
            ws1.cell(row=r, column=c, value=v)
    for r in range(len(samples1) + 2, 52):
        ws1.cell(row=r, column=1, value=r - 1)
    auto_width(ws1, len(h1))

    # --- 怪物区域 ---
    ws2 = wb.create_sheet("怪物区域")
    h2 = [
        "区域等级",
        "区域名称",
        "推荐主角等级",
        "怪物1名称",
        "怪物2名称",
        "BOSS名称",
        "单场经验",
        "单场金币",
        "备注",
    ]
    set_header(ws2, 1, h2)
    for i in range(1, 21):
        r = i + 1
        ws2.cell(row=r, column=1, value=i)
        if i % 5 == 0:
            ws2.cell(row=r, column=9, value="BOSS关")
        if i == 20:
            ws2.cell(row=r, column=9, value="最终BOSS")
    auto_width(ws2, len(h2))

    # --- 装备部位 ---
    ws3 = wb.create_sheet("装备部位")
    h3 = ["部位代码", "中文名", "主要属性倾向", "备注"]
    set_header(ws3, 1, h3)
    slots = [
        ("weapon", "武器", "攻击、攻速", ""),
        ("hat", "帽子", "防御、生命", ""),
        ("clothes", "衣服", "防御、生命", ""),
        ("pants", "裤子", "防御", ""),
        ("ring", "戒指", "攻击、暴击", ""),
        ("bracelet", "手环", "攻速、暴击", ""),
    ]
    for r, row in enumerate(slots, 2):
        for c, v in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=v)
    auto_width(ws3, len(h3))

    # --- 装备等级档 ---
    ws4 = wb.create_sheet("装备等级档")
    h4 = ["装备等级", "适用主角等级", "名称前缀示例", "备注"]
    set_header(ws4, 1, h4)
    tiers = [
        ("1-3", "1-3", "劣质", ""),
        ("4-6", "4-6", "普通", ""),
        ("7-9", "7-9", "精良", ""),
        ("10-12", "10-12", "稀有", ""),
        ("13-15", "13-15", "史诗", ""),
        ("16-20", "16-20", "传说", ""),
        ("21-30", "21-30", "神话", ""),
    ]
    for r, row in enumerate(tiers, 2):
        for c, v in enumerate(row, 1):
            ws4.cell(row=r, column=c, value=v)
    auto_width(ws4, len(h4))

    # --- 装备实例 ---
    ws5 = wb.create_sheet("装备实例")
    h5 = [
        "部位",
        "装备等级",
        "稀有度",
        "装备名称",
        "攻击",
        "防御",
        "生命",
        "暴击%",
        "攻速%",
        "掉落区域等级",
        "掉落率%",
        "商城可售",
    ]
    set_header(ws5, 1, h5)
    rarities = ["普通", "高级", "稀有", "传说", "特制"]
    slots_short = ["weapon", "hat", "clothes", "pants", "ring", "bracelet"]
    row = 2
    for slot in slots_short:
        for rarity in rarities:
            ws5.cell(row=row, column=1, value=slot)
            ws5.cell(row=row, column=2, value=1)
            ws5.cell(row=row, column=3, value=rarity)
            if rarity == "普通":
                ws5.cell(row=row, column=10, value="1-3")
                ws5.cell(row=row, column=11, value=8)
                ws5.cell(row=row, column=12, value="否")
            elif rarity == "高级":
                ws5.cell(row=row, column=10, value="1-3")
                ws5.cell(row=row, column=11, value=3)
                ws5.cell(row=row, column=12, value="是")
            elif rarity == "稀有":
                ws5.cell(row=row, column=10, value="2-4")
                ws5.cell(row=row, column=11, value=1)
                ws5.cell(row=row, column=12, value="否")
            elif rarity == "传说":
                ws5.cell(row=row, column=10, value="3-5")
                ws5.cell(row=row, column=11, value=0.3)
                ws5.cell(row=row, column=12, value="否")
            else:
                ws5.cell(row=row, column=10, value="—")
                ws5.cell(row=row, column=11, value=0.05)
                ws5.cell(row=row, column=12, value="偶尔")
            row += 1
        row += 1
    for extra in range(20):
        ws5.cell(row=row + extra, column=1, value="")
    auto_width(ws5, len(h5))

    # --- 稀有度掉落 ---
    ws6 = wb.create_sheet("稀有度掉落")
    h6 = ["稀有度", "打怪掉落率%", "商城", "备注"]
    set_header(ws6, 1, h6)
    drops = [
        ("普通", 8, "否", ""),
        ("高级", 3, "是（金币）", ""),
        ("稀有", 1, "否", ""),
        ("传说", 0.3, "否", ""),
        ("特制", 0.05, "偶尔刷新", "打怪极低掉落"),
    ]
    for r, row in enumerate(drops, 2):
        for c, v in enumerate(row, 1):
            ws6.cell(row=r, column=c, value=v)
    auto_width(ws6, len(h6))

    # --- 配角 ---
    ws7 = wb.create_sheet("配角")
    h7 = ["配角", "解锁条件", "定位", "备注"]
    set_header(ws7, 1, h7)
    companions = [
        ("群", "默认", "群攻", ""),
        ("粉", "默认", "减速", ""),
        ("编", "金币解锁", "标记破绽", ""),
        ("导", "金币解锁", "高单伤", ""),
    ]
    for r, row in enumerate(companions, 2):
        for c, v in enumerate(row, 1):
            ws7.cell(row=r, column=c, value=v)
    auto_width(ws7, len(h7))

    # --- 命名示例 ---
    ws8 = wb.create_sheet("命名示例")
    ws8.cell(row=1, column=1, value="类型").font = Font(bold=True)
    ws8.cell(row=1, column=2, value="示例").font = Font(bold=True)
    examples = [
        ("区域", "后台更衣室、通告大厅、红毯入口…"),
        ("小怪", "黑粉、水军、代拍、狗仔…"),
        ("BOSS", "毒舌评委、过气顶流、资本大佬…"),
        ("武器", "话筒、剧本、热搜牌…"),
    ]
    for r, (a, b) in enumerate(examples, 2):
        ws8.cell(row=r, column=1, value=a)
        ws8.cell(row=r, column=2, value=b)
    ws8.column_dimensions["A"].width = 12
    ws8.column_dimensions["B"].width = 50

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
